from google.colab import files
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import random

# Patient Name (placeholder)
patient_name = "[patient_name]"

# Define category counts (must sum to 7)
category_a_count = 4  # Number of Category A items
category_b_count = 2  # Number of Category B items
category_c_count = 1  # Number of Category C items

# Verify the sum is 7
total_count = category_a_count + category_b_count + category_c_count
if total_count != 7:
    raise ValueError("Category counts must sum to 7")

# Scale counts to sum to 14 (for 14 days)
days = 14
category_a_quota = round((category_a_count / total_count) * days)  # e.g., 4/7 * 14 ≈ 8
category_b_quota = round((category_b_count / total_count) * days)  # e.g., 2/7 * 14 ≈ 4
category_c_quota = round((category_c_count / total_count) * days)  # e.g., 1/7 * 14 ≈ 2

# Adjust quotas to ensure they sum to 14
total_quota = category_a_quota + category_b_quota + category_c_quota
if total_quota != days:
    diff = days - total_quota
    if category_a_quota >= category_b_quota and category_a_quota >= category_c_quota:
        category_a_quota += diff
    elif category_b_quota >= category_a_quota and category_b_quota >= category_c_quota:
        category_b_quota += diff
    else:
        category_c_quota += diff

# Item list from the document (excluded 'x' items, handled duplicates)
items = [
    {"name": "١-٢ بيضة مسلوقة", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "١-٢ علبة لبن رايب", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "٥ ملاعق جبنة فيتا + خيار", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "٤ قطع جبنة مثلثات", "color": "C", "eat_time": "Breakfast", "group": 1},
    {"name": "قطعة جبنة صفراء كيه٢", "color": "B", "eat_time": "Breakfast", "group": 1},
    {"name": "جبنة قريش + بندورة + خيار", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "١-٢ حبة أفوكادو بدون خبز", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "١ صحن سلطة بدون خبز", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "٥٠ غرام مكسرات (فستق-لوز-عين جمل)", "color": "B", "eat_time": "Breakfast", "group": 1},
    {"name": "٤ ملاعق لبنة كبيرة + ٢ حبة خيار بدون خبز", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "٢ بيضة مقلية بزيت الزيتون + خضار بدون خبز", "color": "A", "eat_time": "Breakfast", "group": 1},
    {"name": "١ صحن متبل باتنجان", "color": "C", "eat_time": "Breakfast", "group": 1},
    {"name": "٣ ملاعق كبيرة حمص + خضار أو سلطة", "color": "C", "eat_time": "Breakfast", "group": 1},
    {"name": "١ تفاحة", "color": "A", "eat_time": "Breakfast", "group": 2},  # Used last entry (A)
    {"name": "خبز شعير ٥٠ جرام بيتا أو فرشوحة", "color": "B", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ حبة خوخ", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ حبة بندورة", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ حبة خيار", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ شريحة بطيخ", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ شريحة شمام", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ حبة أجاص", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "حبة جوافة", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٦ حبات فراولة", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "جريب فروت", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "برتقالة", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "نصف حبة بوملي", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "حبة فرمسون", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢ حبة كيوي", "color": "A", "eat_time": "Breakfast", "group": 2},
    {"name": "٢٥٠ غرام صدر دجاج مشوي", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام سمك أو تونة", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام لحم عجل مشوي بدون دهون", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ علبة تونة مصفاة من الزيت", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن شوربة عدس", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن فول مدمس", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ بازيلاء بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ فاصولياء خضراء بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن سبانخ بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ كوسا مقطعة بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ فول أخضر بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ ملوخية بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ بامية بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن طبيخ يقطين بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كبدة دجاج", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام شاورما بيتية خالي الدهن", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام شيش طاووق", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كباب خالي الدهن", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام طحال مشوي", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام فطر + لحمة + بصل", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام فطر + بيض + بصل", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن لوبياء بدون أرز", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن كوسا مغشي بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن مسقعة بائنجان بدون خبز", "color": "B", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن خبيزة بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن سلق بدون أرز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن زهرة بلبن بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن فول أخضر بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن زهرة بندورة بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن متبل باذنجان مع ٥٠ غرام خبز شعير", "color": "C", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن سماقية بدون خبز", "color": "C", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن رمانية بدون خبز", "color": "C", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كفتة سمك", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام سمك مقلي", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام لحم خروف", "color": "C", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كباب مشوي", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كفتة بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام تايلندي + خضار بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام شاورما بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام جناحين مشوية", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام فخد دجاج", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام كبدة دجاج أو عجل بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام طحال أو فشة بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام حبش", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "٢٥٠ غرام ستيكات دجاج", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "علبة فطر مع بيض و بصل", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "١ صحن ملوخية ورق بدون خبز", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "صحن طبيخ حماصيص", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "صحن طبيخ رجلة أو بقلة", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "صحن طبيخ سلق و عدس", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "شوربة خضار", "color": "A", "eat_time": "Lunch", "group": 1},
    {"name": "سلطة خضراء بدون ذرة", "color": "A", "eat_time": "Lunch", "group": 2},
    {"name": "شوربة خضار بدون بطاطس أو ذرة", "color": "A", "eat_time": "Lunch", "group": 2},
    {"name": "شوربة ملفوف", "color": "A", "eat_time": "Lunch", "group": 2},
    {"name": "شوربة فطر و بصل بدون كريمة", "color": "A", "eat_time": "Lunch", "group": 2},
    {"name": "خضار بدون سلطة أو شوربة", "color": "A", "eat_time": "Lunch", "group": 2},
    {"name": "١ صحن بائنجان مقلي", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن متبل باذنجان", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن خيار + لبن", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن سلطة ملفوف مع مايونيز أو خل أو زيت", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن سلطة خضراء + طحينة", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن شكشوكة بدون خبز", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ جرام فستق حلبي", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ جرام فستق سوداني", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام لوز", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام بندق", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام كاجو", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام عين جمل", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام بذر عين شمس", "color": "B", "eat_time": "Dinner", "group": 1},
    {"name": "صحن عجة بيض + بقدونس + بصل", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "٣ أو ٤ حبات فاكهة صنف واحد (تفاح، مشمس، كيوي، أجاص، خوخ)", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "علبة فول مدمس", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "٢ شريحة بطيخ", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "٢ شريحة شمام", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام عنب", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "٢ حبة تين", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١٠٠ غرام مانجا", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "حبة جوافة", "color": "A", "eat_time": "Dinner", "group": 1},
    {"name": "١ صحن سلطة يونانية (سلطة خضراء + جبنة بيضاء قليلة الدسم)", "color": "A", "eat_time": "Dinner", "group": 1}
]

# Days of the week in Arabic (two weeks)
days = [
    "السبت", "الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة",
    "السبت", "الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة"
]

# Initialize counters for category quotas per meal type
category_counts = {
    "Breakfast": {"A": 0, "B": 0, "C": 0},
    "Lunch": {"A": 0, "B": 0, "C": 0},
    "Dinner": {"A": 0, "B": 0, "C": 0}
}

# Function to select items respecting category quotas and no consecutive repeats
def select_meal_items(meal_type, group, used_items, prev_day_items, category_counts, quotas):
    available_items = [item for item in items if item["eat_time"] == meal_type and item["group"] == group]
    
    if group == 1:
        available_items = [item for item in available_items if item["name"] not in prev_day_items]
        
        filtered_items = []
        for item in available_items:
            category = item["color"]
            if category == "A" and category_counts[meal_type]["A"] < quotas["A"]:
                filtered_items.append(item)
            elif category == "B" and category_counts[meal_type]["B"] < quotas["B"]:
                filtered_items.append(item)
            elif category == "C" and category_counts[meal_type]["C"] < quotas["C"]:
                filtered_items.append(item)
        available_items = filtered_items
    
    if not available_items:
        available_items = [item for item in items if item["eat_time"] == meal_type and item["group"] == group]
        available_items = [item for item in available_items if item["name"] not in prev_day_items]
    
    if available_items:
        selected_item = random.choice(available_items)
        if group == 1:
            category_counts[meal_type][selected_item["color"]] += 1
        return selected_item["name"]
    return None

# Function to get combined Group 1 + Group 2 alternatives for Breakfast and Lunch
def get_combined_alternatives(meal_type, selected_combination):
    group1_items = [item["name"] for item in items if item["eat_time"] == meal_type and item["group"] == 1]
    group2_items = [item["name"] for item in items if item["eat_time"] == meal_type and item["group"] == 2]
    
    all_combinations = [f"{g1} + {g2}" for g1 in group1_items for g2 in group2_items]
    alternatives = [combo for combo in all_combinations if combo != selected_combination]
    
    random.shuffle(alternatives)
    # return alternatives[:10]  # 10 alternatives as specified
    return alternatives

# Function to get Group 1 alternatives for Dinner
def get_dinner_alternatives(meal_type, selected_item):
    group1_items = [item["name"] for item in items if item["eat_time"] == meal_type and item["group"] == 1]
    alternatives = [item for item in group1_items if item != selected_item]
    random.shuffle(alternatives)
    # return alternatives[:10]  # 10 alternatives as specified
    return alternatives

# Generate meal plan
meal_plan = []
used_breakfast_items = []
used_lunch_items = []
breakfast_alternatives = []
lunch_alternatives = []
dinner_alternatives = []

quotas = {
    "A": category_a_quota,
    "B": category_b_quota,
    "C": category_c_quota
}

for i, day in enumerate(days):
    day_plan = {"day": day}
    
    # Breakfast: Group 1 + Group 2
    prev_breakfast_items = used_breakfast_items[-1] if used_breakfast_items else []
    breakfast_group1 = select_meal_items("Breakfast", 1, used_breakfast_items, prev_breakfast_items, category_counts, quotas)
    breakfast_group2 = select_meal_items("Breakfast", 2, [], [], category_counts, quotas)
    breakfast_combo = f"{breakfast_group1} + {breakfast_group2}"
    day_plan["breakfast"] = breakfast_combo
    breakfast_alts = get_combined_alternatives("Breakfast", breakfast_combo)
    breakfast_alternatives.append(breakfast_alts)
    used_breakfast_items.append([breakfast_group1])
    
    # Lunch: Group 1 + Group 2
    prev_lunch_items = used_lunch_items[-1] if used_lunch_items else []
    lunch_group1 = select_meal_items("Lunch", 1, used_lunch_items, prev_lunch_items, category_counts, quotas)
    lunch_group2 = select_meal_items("Lunch", 2, [], [], category_counts, quotas)
    lunch_combo = f"{lunch_group1} + {lunch_group2}"
    day_plan["lunch"] = lunch_combo
    lunch_alts = get_combined_alternatives("Lunch", lunch_combo)
    lunch_alternatives.append(lunch_alts)
    used_lunch_items.append([lunch_group1])
    
    # Dinner: Single Group 1 item
    dinner_item = select_meal_items("Dinner", 1, [], [], category_counts, quotas)
    day_plan["dinner"] = dinner_item
    dinner_alts = get_dinner_alternatives("Dinner", dinner_item)
    dinner_alternatives.append(dinner_alts)
    
    meal_plan.append(day_plan)

# Create Excel file
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "خطة الوجبات"  # Modified to remove patient name

# Write headers
headers = ["اليوم", "الإفطار", "الغداء", "العشاء"]
for col, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col).value = header

# Write meal plan
for row, day_plan in enumerate(meal_plan, 2):
    sheet.cell(row=row, column=1).value = day_plan["day"]
    sheet.cell(row=row, column=2).value = day_plan["breakfast"]
    sheet.cell(row=row, column=3).value = day_plan["lunch"]
    sheet.cell(row=row, column=4).value = day_plan["dinner"]

# Create sheets for alternatives
breakfast_alt_sheet = workbook.create_sheet("Breakfast_Alternatives")
lunch_alt_sheet = workbook.create_sheet("Lunch_Alternatives")
dinner_alt_sheet = workbook.create_sheet("Dinner_Alternatives")

# Write breakfast alternatives (combined Group 1 + Group 2)
for row, alts in enumerate(breakfast_alternatives, 1):
    for col, alt in enumerate(alts, 1):
        breakfast_alt_sheet.cell(row=row, column=col).value = alt

# Write lunch alternatives (combined Group 1 + Group 2)
for row, alts in enumerate(lunch_alternatives, 1):
    for col, alt in enumerate(alts, 1):
        lunch_alt_sheet.cell(row=row, column=col).value = alt

# Write dinner alternatives (Group 1 only)
for row, alts in enumerate(dinner_alternatives, 1):
    for col, alt in enumerate(alts, 1):
        dinner_alt_sheet.cell(row=row, column=col).value = alt

# Add Data Validation for alternatives
for row in range(2, len(meal_plan) + 2):
    dv_breakfast = DataValidation(
        type="list",
        formula1=f"Breakfast_Alternatives!A{row-1}:J{row-1}",  # 10 columns (A:J)
        allow_blank=True
    )
    dv_breakfast.add(f"B{row}")
    sheet.add_data_validation(dv_breakfast)
    
    dv_lunch = DataValidation(
        type="list",
        formula1=f"Lunch_Alternatives!A{row-1}:J{row-1}",  # 10 columns (A:J)
        allow_blank=True
    )
    dv_lunch.add(f"C{row}")
    sheet.add_data_validation(dv_lunch)
    
    dv_dinner = DataValidation(
        type="list",
        formula1=f"Dinner_Alternatives!A{row-1}:J{row-1}",  # 10 columns (A:J)
        allow_blank=True
    )
    dv_dinner.add(f"D{row}")
    sheet.add_data_validation(dv_dinner)

# Save and download Excel file
workbook.save("meal_plan.xlsx")
files.download("meal_plan.xlsx")

# Print category counts for verification
print("Breakfast Category Counts:", category_counts["Breakfast"])
print("Lunch Category Counts:", category_counts["Lunch"])
print("Dinner Category Counts:", category_counts["Dinner"])